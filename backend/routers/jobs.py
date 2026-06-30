from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional
import io
import os
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.platypus import Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

from database import get_connection, dict_cursor
from models import (
    JobCreate, JobUpdate, JobResponse,
    CalculationPreview, CalculationResult,
    RepeatOrderCreate, RepeatOrderResponse,
    BulkPaperPlannedRequest,
)
from calculations import calculate_job

router = APIRouter(prefix="/api/jobs", tags=["jobs"])

# ── Logo ──────────────────────────────────────────────────────────────────────

LOGO_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets", "logo.png")
_logo_bytes: Optional[bytes] = None
_logo_loaded = False


def _get_logo() -> Optional[bytes]:
    global _logo_bytes, _logo_loaded
    if _logo_loaded:
        return _logo_bytes
    _logo_loaded = True
    try:
        with open(LOGO_PATH, "rb") as f:
            _logo_bytes = f.read()
    except Exception:
        _logo_bytes = None
    return _logo_bytes


# ── Search / sort helpers ─────────────────────────────────────────────────────

_SEARCH_FILTER = """
    customer_name                      ILIKE %s OR
    job_name                           ILIKE %s OR
    artworks                           ILIKE %s OR
    CAST(length          AS TEXT)      ILIKE %s OR
    CAST(width           AS TEXT)      ILIKE %s OR
    CAST(height          AS TEXT)      ILIKE %s OR
    CAST(gsm             AS TEXT)      ILIKE %s OR
    paper_quality                      ILIKE %s OR
    CAST(order_quantity  AS TEXT)      ILIKE %s OR
    CAST(sheet_length    AS TEXT)      ILIKE %s OR
    CAST(sheet_width     AS TEXT)      ILIKE %s OR
    CAST(ups             AS TEXT)      ILIKE %s OR
    CAST(final_sheets    AS TEXT)      ILIKE %s OR
    CAST(total_kg        AS TEXT)      ILIKE %s OR
    TO_CHAR(created_at, 'DD/MM/YYYY') ILIKE %s
"""

_SORT_MAP = {
    "newest":      "created_at DESC",
    "oldest":      "created_at ASC",
    "customer_az": "customer_name ASC",
    "customer_za": "customer_name DESC",
    "job_az":      "job_name ASC",
    "job_za":      "job_name DESC",
    "order_qty":   "order_quantity DESC",
    "sheets":      "final_sheets DESC",
    "kg":          "total_kg DESC",
}


def _fmt_box(job: dict) -> str:
    parts = [str(job[k]) for k in ("length", "width", "height") if job.get(k) is not None]
    return "×".join(parts) if parts else "—"


def _parse_date(label: str, value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid {label}: expected YYYY-MM-DD")


# ── Repeat-order enrichment helpers ──────────────────────────────────────────

def _enrich_rows(cur, rows: list[dict]) -> list[dict]:
    """Attach repeat_order_count and repeat_total_qty to each job dict."""
    if not rows:
        return rows
    job_ids = [r["id"] for r in rows]
    placeholders = ",".join(["%s"] * len(job_ids))
    cur.execute(
        f"""
        SELECT job_id,
               COUNT(*)                          AS cnt,
               COALESCE(SUM(order_quantity), 0)  AS total_qty
        FROM repeat_orders
        WHERE job_id IN ({placeholders})
        GROUP BY job_id
        """,
        job_ids,
    )
    stats = {r["job_id"]: dict(r) for r in cur.fetchall()}
    for row in rows:
        s = stats.get(row["id"], {})
        row["repeat_order_count"] = int(s.get("cnt", 0))
        row["repeat_total_qty"]   = int(s.get("total_qty", 0))
    return rows


def _enrich_one(cur, row: dict) -> dict:
    cur.execute(
        """
        SELECT COUNT(*) AS cnt, COALESCE(SUM(order_quantity), 0) AS total_qty
        FROM repeat_orders WHERE job_id = %s
        """,
        [row["id"]],
    )
    s = dict(cur.fetchone())
    row["repeat_order_count"] = int(s["cnt"])
    row["repeat_total_qty"]   = int(s["total_qty"])
    return row


# ── Job fetch helpers ─────────────────────────────────────────────────────────

def _fetch_jobs(
    search: Optional[str],
    sort_by: Optional[str],
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
) -> list[dict]:
    order_clause = _SORT_MAP.get(sort_by or "newest", "created_at DESC")
    conn = get_connection()
    cur  = dict_cursor(conn)

    conditions: list[str] = []
    params: list = []

    if search:
        conditions.append(f"({_SEARCH_FILTER})")
        params.extend([f"%{search}%"] * 15)

    if from_date and to_date:
        conditions.append("created_at::date BETWEEN %s AND %s")
        params.extend([from_date, to_date])

    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    # paper_planned=FALSE (0) sorts before TRUE (1) — planned jobs go to bottom
    cur.execute(
        f"SELECT * FROM jobs {where_clause} ORDER BY paper_planned ASC, {order_clause}",
        params,
    )
    rows = [dict(r) for r in cur.fetchall()]
    _enrich_rows(cur, rows)
    conn.close()
    return rows


def _fetch_jobs_by_ids(job_ids: list[int]) -> list[dict]:
    if not job_ids:
        return []
    conn = get_connection()
    cur  = dict_cursor(conn)
    placeholders = ",".join(["%s"] * len(job_ids))
    cur.execute(
        f"SELECT * FROM jobs WHERE id IN ({placeholders}) ORDER BY created_at DESC",
        job_ids,
    )
    rows = [dict(r) for r in cur.fetchall()]
    _enrich_rows(cur, rows)
    conn.close()
    return rows


def _fetch_jobs_by_ids_with_repeats(job_ids: list[int]) -> list[dict]:
    """Fetch jobs AND their full repeat_orders list — used by supplier export."""
    if not job_ids:
        return []
    conn = get_connection()
    cur  = dict_cursor(conn)
    placeholders = ",".join(["%s"] * len(job_ids))
    cur.execute(
        f"SELECT * FROM jobs WHERE id IN ({placeholders}) ORDER BY created_at DESC",
        job_ids,
    )
    rows = [dict(r) for r in cur.fetchall()]

    if rows:
        cur.execute(
            f"""
            SELECT * FROM repeat_orders
            WHERE job_id IN ({placeholders})
            ORDER BY created_at ASC
            """,
            job_ids,
        )
        repeat_rows = [dict(r) for r in cur.fetchall()]
        by_job: dict[int, list] = {}
        for r in repeat_rows:
            by_job.setdefault(r["job_id"], []).append(r)
        for row in rows:
            row["repeat_orders"] = by_job.get(row["id"], [])

    conn.close()
    return rows


def _parse_job_ids(ids: str) -> list[int]:
    try:
        job_ids = [int(i.strip()) for i in ids.split(",") if i.strip().isdigit()]
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid job IDs.")
    if not job_ids:
        raise HTTPException(status_code=400, detail="No valid job IDs provided.")
    return job_ids


# ── Calculation preview ───────────────────────────────────────────────────────

@router.post("/calculate", response_model=CalculationResult)
def calculate_preview(data: CalculationPreview):
    return calculate_job(
        order_quantity=data.order_quantity,
        ups=data.ups,
        sheet_length=data.sheet_length,
        sheet_width=data.sheet_width,
        gsm=data.gsm,
    )


# ── Company / GSM autocomplete ────────────────────────────────────────────────

@router.get("/companies", response_model=list[str])
def get_companies():
    conn = get_connection()
    cur  = conn.cursor()
    cur.execute("SELECT DISTINCT customer_name FROM jobs ORDER BY customer_name")
    rows = cur.fetchall()
    conn.close()
    return [r[0] for r in rows]


@router.get("/gsm-values", response_model=list[int])
def get_gsm_values():
    conn = get_connection()
    cur  = conn.cursor()
    cur.execute("SELECT DISTINCT gsm FROM jobs WHERE gsm IS NOT NULL ORDER BY gsm")
    rows = cur.fetchall()
    conn.close()
    return [r[0] for r in rows]


# ── Bulk paper-planned (must be before /{job_id}) ────────────────────────────

@router.post("/bulk/paper-planned", response_model=list[JobResponse])
def bulk_paper_planned(data: BulkPaperPlannedRequest):
    if not data.job_ids:
        return []
    conn = get_connection()
    cur  = dict_cursor(conn)
    try:
        placeholders = ",".join(["%s"] * len(data.job_ids))
        cur.execute(
            f"""
            UPDATE jobs SET paper_planned = %s, updated_at = NOW()
            WHERE id IN ({placeholders})
            RETURNING *
            """,
            [data.planned] + data.job_ids,
        )
        rows = [dict(r) for r in cur.fetchall()]
        conn.commit()
        _enrich_rows(cur, rows)
        return [JobResponse(**r) for r in rows]
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ── Full-export builders ──────────────────────────────────────────────────────

def _build_jobs_excel(rows: list[dict]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Records"

    thin   = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    logo_bytes      = _get_logo()
    logo_row_offset = 0
    if logo_bytes:
        try:
            img        = XLImage(io.BytesIO(logo_bytes))
            img.width  = 140
            img.height = 70
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 40
            ws.row_dimensions[2].height = 36
            logo_row_offset = 2
        except Exception:
            logo_row_offset = 0

    title_row = logo_row_offset + 1
    hdr_row   = logo_row_offset + 2

    last_col = get_column_letter(18)
    ws.merge_cells(f"A{title_row}:{last_col}{title_row}")
    tc           = ws.cell(row=title_row, column=1)
    tc.value     = "Shri Neminath Printers & Packaging — Job Record Register"
    tc.font      = Font(bold=True, size=13, color="FFFFFF")
    tc.fill      = PatternFill("solid", fgColor="0F172A")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[title_row].height = 32

    headers = [
        "Date Created", "Customer Name", "Job Name", "Artworks",
        "L (cm)", "W (cm)", "H (cm)",
        "GSM", "Paper Quality", "Order Qty",
        "Sheet L (cm)", "Sheet W (cm)", "UPS",
        "Base Sheets", "Wastage %", "Final Sheets", "Total KG",
        "Last Modified",
    ]
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="1E40AF")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, h in enumerate(headers, 1):
        c           = ws.cell(row=hdr_row, column=col, value=h)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = hdr_align
        c.border    = border
    ws.row_dimensions[hdr_row].height = 36

    alt_fill   = PatternFill("solid", fgColor="EFF6FF")
    data_start = logo_row_offset + 3
    for ri, job in enumerate(rows, data_start):
        fill = alt_fill if ri % 2 == 0 else None
        vals = [
            job["created_at"].strftime("%d/%m/%Y") if job.get("created_at") else "",
            job["customer_name"],
            job["job_name"],
            job.get("artworks", ""),
            job.get("length", ""),
            job.get("width", ""),
            job.get("height", ""),
            job["gsm"],
            job["paper_quality"],
            job["order_quantity"],
            job["sheet_length"],
            job["sheet_width"],
            job["ups"],
            job["base_sheets"],
            f"{job['wastage_percentage']}%",
            job["final_sheets"],
            job["total_kg"],
            job["updated_at"].strftime("%d/%m/%Y %H:%M") if job.get("updated_at") else "",
        ]
        for col, v in enumerate(vals, 1):
            c           = ws.cell(row=ri, column=col, value=v)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border
            if fill:
                c.fill = fill

    col_widths = [14, 22, 22, 20, 9, 9, 9, 8, 18, 11, 12, 12, 8, 13, 11, 13, 11, 18]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _build_jobs_pdf(rows: list[dict]) -> io.BytesIO:
    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=landscape(A4),
        rightMargin=0.7 * cm, leftMargin=0.7 * cm,
        topMargin=1.5 * cm,  bottomMargin=1 * cm,
    )
    styles      = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "AppTitle", parent=styles["Heading1"],
        fontSize=13, textColor=colors.HexColor("#0F172A"),
        spaceAfter=2, alignment=1,
    )
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        fontSize=8, textColor=colors.HexColor("#64748B"),
        spaceAfter=10, alignment=1,
    )

    elements   = []
    logo_bytes = _get_logo()
    title_para = Paragraph("Shri Neminath Printers &amp; Packaging", title_style)
    sub_para   = Paragraph(
        f"Job Record Register &nbsp;|&nbsp; Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        f" &nbsp;|&nbsp; Total Records: {len(rows)}",
        sub_style,
    )

    if logo_bytes:
        try:
            logo_img = RLImage(io.BytesIO(logo_bytes))
            logo_img._restrictSize(3 * cm, 2 * cm)
            header_data = [[logo_img, [title_para, sub_para]]]
            header_tbl  = Table(header_data, colWidths=[3.5 * cm, None])
            header_tbl.setStyle(TableStyle([
                ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING",  (0, 0), (0, 0),  0),
                ("RIGHTPADDING", (0, 0), (0, 0),  8),
                ("LEFTPADDING",  (1, 0), (1, 0),  8),
            ]))
            elements.append(header_tbl)
        except Exception:
            elements += [title_para, sub_para]
    else:
        elements += [title_para, sub_para]

    tbl_data = [[
        "Date", "Customer", "Job Name", "Artworks", "Box (L×W×H)", "GSM",
        "Quality", "Qty", "Sheet (L×W)", "UPS", "Sheets", "KG",
    ]]
    for job in rows:
        tbl_data.append([
            job["created_at"].strftime("%d/%m/%y") if job.get("created_at") else "",
            job["customer_name"],
            job["job_name"],
            job.get("artworks", ""),
            _fmt_box(job),
            str(job["gsm"]),
            job["paper_quality"],
            str(job["order_quantity"]),
            f"{job['sheet_length']}×{job['sheet_width']}",
            str(job["ups"]),
            str(job["final_sheets"]),
            str(job["total_kg"]),
        ])

    col_w = [1.8*cm, 3.2*cm, 3.2*cm, 3*cm, 3*cm, 1.3*cm, 2.8*cm, 1.8*cm, 2.5*cm, 1.3*cm, 2.2*cm, 1.8*cm]
    tbl   = Table(tbl_data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0),  colors.HexColor("#1E40AF")),
        ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",       (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",       (0, 0), (-1, 0),  7),
        ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF6FF")]),
        ("FONTSIZE",       (0, 1), (-1, -1), 6.5),
        ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("TOPPADDING",     (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 3),
    ]))
    elements.append(tbl)
    doc.build(elements)
    out.seek(0)
    return out


# ── Date-range exports ────────────────────────────────────────────────────────

@router.get("/export/excel")
def export_excel(
    search: Optional[str] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
):
    rows = _fetch_jobs(search, "newest", _parse_date("from_date", from_date), _parse_date("to_date", to_date))
    out  = _build_jobs_excel(rows)
    fname = f"job_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/export/pdf")
def export_pdf(
    search: Optional[str] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
):
    rows  = _fetch_jobs(search, "newest", _parse_date("from_date", from_date), _parse_date("to_date", to_date))
    out   = _build_jobs_pdf(rows)
    fname = f"job_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Selected-jobs exports (full normal format) ────────────────────────────────

@router.get("/export/selected/excel")
def export_selected_excel(ids: str = Query(...)):
    rows  = _fetch_jobs_by_ids(_parse_job_ids(ids))
    out   = _build_jobs_excel(rows)
    fname = f"job_records_selected_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/export/selected/pdf")
def export_selected_pdf(ids: str = Query(...)):
    rows  = _fetch_jobs_by_ids(_parse_job_ids(ids))
    out   = _build_jobs_pdf(rows)
    fname = f"job_records_selected_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Supplier export builders (with order history) ─────────────────────────────

def _build_supplier_excel(rows: list[dict]) -> io.BytesIO:
    """
    rows must have a 'repeat_orders' list on each job dict
    (from _fetch_jobs_by_ids_with_repeats).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supplier Sheet"

    thin   = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    logo_bytes      = _get_logo()
    logo_row_offset = 0
    if logo_bytes:
        try:
            img        = XLImage(io.BytesIO(logo_bytes))
            img.width  = 100
            img.height = 50
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 26
            logo_row_offset = 2
        except Exception:
            logo_row_offset = 0

    title_row  = logo_row_offset + 1
    sub_row    = logo_row_offset + 2
    hdr_row    = logo_row_offset + 3
    data_start = hdr_row + 1

    last_col = get_column_letter(10)
    ws.merge_cells(f"A{title_row}:{last_col}{title_row}")
    tc           = ws.cell(row=title_row, column=1)
    tc.value     = "Shri Neminath Printers & Packaging — Supplier Paper Requirement Sheet"
    tc.font      = Font(bold=True, size=12, color="FFFFFF")
    tc.fill      = PatternFill("solid", fgColor="0F172A")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[title_row].height = 28

    ws.merge_cells(f"A{sub_row}:{last_col}{sub_row}")
    sc           = ws.cell(row=sub_row, column=1)
    sc.value     = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Jobs: {len(rows)}"
    sc.font      = Font(size=9, color="94A3B8")
    sc.fill      = PatternFill("solid", fgColor="0F172A")
    sc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[sub_row].height = 18

    headers = [
        "#", "Customer / Job", "Paper Quality", "GSM",
        "Sheet L (cm)", "Sheet W (cm)",
        "Order Date", "Order Qty", "Final Sheets", "Total KG",
    ]
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="1E40AF")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, h in enumerate(headers, 1):
        c           = ws.cell(row=hdr_row, column=col, value=h)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = hdr_align
        c.border    = border
    ws.row_dimensions[hdr_row].height = 32

    alt_fill   = PatternFill("solid", fgColor="EFF6FF")
    total_fill = PatternFill("solid", fgColor="D1FAE5")   # light green for totals
    repeat_fill = PatternFill("solid", fgColor="F0FDF4")  # very light green for repeat rows

    ri = data_start
    for job_idx, job in enumerate(rows):
        repeat_orders = job.get("repeat_orders", [])
        has_repeats   = len(repeat_orders) > 0
        base_fill     = alt_fill if job_idx % 2 == 0 else None

        # ── Original order row ───────────────────────────────────────────────
        orig_date = job["created_at"].strftime("%d %b %Y") if job.get("created_at") else ""
        vals = [
            job_idx + 1,
            f"{job['customer_name']} / {job['job_name']}",
            job["paper_quality"],
            job["gsm"],
            job["sheet_length"],
            job["sheet_width"],
            orig_date,
            job["order_quantity"],
            job["final_sheets"],
            job["total_kg"],
        ]
        for col, v in enumerate(vals, 1):
            c           = ws.cell(row=ri, column=col, value=v)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border
            if base_fill:
                c.fill = base_fill
        ws.row_dimensions[ri].height = 20
        ri += 1

        # ── Repeat order rows ────────────────────────────────────────────────
        for rep in repeat_orders:
            calc = calculate_job(
                order_quantity=rep["order_quantity"],
                ups=job["ups"],
                sheet_length=job["sheet_length"],
                sheet_width=job["sheet_width"],
                gsm=job["gsm"],
            )
            rep_date = rep["created_at"].strftime("%d %b %Y") if rep.get("created_at") else ""
            rep_vals = [
                "",
                "  ↳ Repeat Order",
                "", "", "", "",
                rep_date,
                rep["order_quantity"],
                calc["final_sheets"],
                calc["total_kg"],
            ]
            for col, v in enumerate(rep_vals, 1):
                c           = ws.cell(row=ri, column=col, value=v)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = border
                c.fill      = repeat_fill
                if col == 2:
                    c.font  = Font(italic=True, color="166534", size=9)
                    c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[ri].height = 18
            ri += 1

        # ── TOTAL row (only when there are repeats) ──────────────────────────
        if has_repeats:
            total_qty = job["order_quantity"] + sum(r["order_quantity"] for r in repeat_orders)
            total_calc = calculate_job(
                order_quantity=total_qty,
                ups=job["ups"],
                sheet_length=job["sheet_length"],
                sheet_width=job["sheet_width"],
                gsm=job["gsm"],
            )
            total_vals = [
                "", "TOTAL", "", "", "", "",
                "",
                total_qty,
                total_calc["final_sheets"],
                total_calc["total_kg"],
            ]
            for col, v in enumerate(total_vals, 1):
                c           = ws.cell(row=ri, column=col, value=v)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = border
                c.fill      = total_fill
                c.font      = Font(bold=True, color="166534", size=10)
                if col == 2:
                    c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[ri].height = 22
            ri += 1

    col_widths = [5, 32, 18, 8, 13, 13, 14, 12, 13, 11]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _build_supplier_pdf(rows: list[dict]) -> io.BytesIO:
    """
    rows must have a 'repeat_orders' list on each job dict.
    """
    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=landscape(A4),
        rightMargin=0.8 * cm, leftMargin=0.8 * cm,
        topMargin=1.5 * cm, bottomMargin=1 * cm,
    )
    styles      = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "SupTitle", parent=styles["Heading1"],
        fontSize=12, textColor=colors.HexColor("#0F172A"),
        spaceAfter=2, alignment=1,
    )
    sub_style = ParagraphStyle(
        "SupSub", parent=styles["Normal"],
        fontSize=8, textColor=colors.HexColor("#64748B"),
        spaceAfter=10, alignment=1,
    )

    elements   = []
    logo_bytes = _get_logo()
    title_para = Paragraph("Shri Neminath Printers &amp; Packaging", title_style)
    sub_para   = Paragraph(
        f"Supplier Paper Requirement Sheet &nbsp;|&nbsp; "
        f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        f" &nbsp;|&nbsp; Jobs: {len(rows)}",
        sub_style,
    )

    if logo_bytes:
        try:
            logo_img = RLImage(io.BytesIO(logo_bytes))
            logo_img._restrictSize(2.5 * cm, 1.8 * cm)
            header_data = [[logo_img, [title_para, sub_para]]]
            header_tbl  = Table(header_data, colWidths=[3 * cm, None])
            header_tbl.setStyle(TableStyle([
                ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING",  (0, 0), (0, 0),  0),
                ("RIGHTPADDING", (0, 0), (0, 0),  8),
                ("LEFTPADDING",  (1, 0), (1, 0),  8),
            ]))
            elements.append(header_tbl)
        except Exception:
            elements += [title_para, sub_para]
    else:
        elements += [title_para, sub_para]

    tbl_data = [[
        "#", "Customer / Job", "Paper Quality", "GSM",
        "Sheet L", "Sheet W", "Order Date", "Order Qty", "Sheets", "KG",
    ]]

    green_rows: list[int] = []   # 0-based data rows that are TOTAL rows

    data_row_idx = 1
    for job_idx, job in enumerate(rows):
        repeat_orders = job.get("repeat_orders", [])
        has_repeats   = len(repeat_orders) > 0

        orig_date = job["created_at"].strftime("%d %b %y") if job.get("created_at") else ""
        tbl_data.append([
            str(job_idx + 1),
            f"{job['customer_name']} / {job['job_name']}",
            job["paper_quality"],
            str(job["gsm"]),
            str(job["sheet_length"]),
            str(job["sheet_width"]),
            orig_date,
            str(job["order_quantity"]),
            str(job["final_sheets"]),
            str(job["total_kg"]),
        ])
        data_row_idx += 1

        for rep in repeat_orders:
            calc = calculate_job(
                order_quantity=rep["order_quantity"],
                ups=job["ups"],
                sheet_length=job["sheet_length"],
                sheet_width=job["sheet_width"],
                gsm=job["gsm"],
            )
            rep_date = rep["created_at"].strftime("%d %b %y") if rep.get("created_at") else ""
            tbl_data.append([
                "",
                "  ↳ Repeat Order",
                "", "", "", "",
                rep_date,
                str(rep["order_quantity"]),
                str(calc["final_sheets"]),
                str(calc["total_kg"]),
            ])
            data_row_idx += 1

        if has_repeats:
            total_qty = job["order_quantity"] + sum(r["order_quantity"] for r in repeat_orders)
            total_calc = calculate_job(
                order_quantity=total_qty,
                ups=job["ups"],
                sheet_length=job["sheet_length"],
                sheet_width=job["sheet_width"],
                gsm=job["gsm"],
            )
            tbl_data.append([
                "", "TOTAL", "", "", "", "", "",
                str(total_qty),
                str(total_calc["final_sheets"]),
                str(total_calc["total_kg"]),
            ])
            green_rows.append(data_row_idx)
            data_row_idx += 1

    col_w = [0.8*cm, 5*cm, 3*cm, 1.2*cm, 1.8*cm, 1.8*cm, 2*cm, 2.2*cm, 2*cm, 1.8*cm]
    tbl   = Table(tbl_data, colWidths=col_w, repeatRows=1)

    style_cmds = [
        ("BACKGROUND",     (0, 0), (-1, 0),  colors.HexColor("#1E40AF")),
        ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",       (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",       (0, 0), (-1, 0),  7),
        ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF6FF")]),
        ("FONTSIZE",       (0, 1), (-1, -1), 6.5),
        ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("TOPPADDING",     (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 3),
    ]
    for gr in green_rows:
        style_cmds += [
            ("BACKGROUND", (0, gr), (-1, gr), colors.HexColor("#D1FAE5")),
            ("FONTNAME",   (0, gr), (-1, gr), "Helvetica-Bold"),
            ("TEXTCOLOR",  (0, gr), (-1, gr), colors.HexColor("#166534")),
        ]

    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)
    doc.build(elements)
    out.seek(0)
    return out


# ── Supplier export endpoints ─────────────────────────────────────────────────

@router.get("/export/supplier/excel")
def export_supplier_excel(ids: str = Query(...)):
    rows  = _fetch_jobs_by_ids_with_repeats(_parse_job_ids(ids))
    out   = _build_supplier_excel(rows)
    fname = f"supplier_sheet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/export/supplier/pdf")
def export_supplier_pdf(ids: str = Query(...)):
    rows  = _fetch_jobs_by_ids_with_repeats(_parse_job_ids(ids))
    out   = _build_supplier_pdf(rows)
    fname = f"supplier_sheet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── CRUD ──────────────────────────────────────────────────────────────────────

@router.get("", response_model=list[JobResponse])
def get_jobs(
    search:  Optional[str] = Query(None),
    sort_by: Optional[str] = Query("newest"),
):
    try:
        rows = _fetch_jobs(search, sort_by)
        return [JobResponse(**r) for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("", response_model=JobResponse, status_code=201)
def create_job(job: JobCreate):
    conn = get_connection()
    cur  = dict_cursor(conn)
    try:
        calc = calculate_job(
            order_quantity=job.order_quantity,
            ups=job.ups,
            sheet_length=job.sheet_length,
            sheet_width=job.sheet_width,
            gsm=job.gsm,
        )
        cur.execute("""
            INSERT INTO jobs (
                customer_name, job_name, artworks,
                length, width, height,
                gsm, paper_quality,
                order_quantity, sheet_length, sheet_width,
                ups, printing_type, remarks,
                base_sheets, wastage_percentage, final_sheets, total_kg
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING *
        """, [
            job.customer_name, job.job_name, job.artworks,
            job.length, job.width, job.height,
            job.gsm, job.paper_quality,
            job.order_quantity,
            job.sheet_length, job.sheet_width,
            job.ups, job.printing_type, job.remarks,
            calc["base_sheets"], calc["wastage_percentage"],
            calc["final_sheets"], calc["total_kg"],
        ])
        row = dict(cur.fetchone())
        conn.commit()
        _enrich_one(cur, row)
        return JobResponse(**row)
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ── Repeat orders ─────────────────────────────────────────────────────────────

@router.get("/{job_id}/repeat-orders", response_model=list[RepeatOrderResponse])
def get_repeat_orders(job_id: int):
    conn = get_connection()
    cur  = dict_cursor(conn)
    cur.execute("SELECT id FROM jobs WHERE id = %s", [job_id])
    if not cur.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found.")
    cur.execute(
        "SELECT * FROM repeat_orders WHERE job_id = %s ORDER BY created_at ASC",
        [job_id],
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return [RepeatOrderResponse(**r) for r in rows]


@router.post("/{job_id}/repeat-orders", response_model=RepeatOrderResponse, status_code=201)
def create_repeat_order(job_id: int, data: RepeatOrderCreate):
    conn = get_connection()
    cur  = dict_cursor(conn)
    try:
        cur.execute("SELECT id FROM jobs WHERE id = %s", [job_id])
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Job not found.")
        cur.execute(
            """
            INSERT INTO repeat_orders (job_id, order_quantity, remarks)
            VALUES (%s, %s, %s)
            RETURNING *
            """,
            [job_id, data.order_quantity, data.remarks],
        )
        row = dict(cur.fetchone())
        conn.commit()
        return RepeatOrderResponse(**row)
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ── Paper-planned toggle ──────────────────────────────────────────────────────

@router.patch("/{job_id}/paper-planned", response_model=JobResponse)
def toggle_paper_planned(job_id: int):
    conn = get_connection()
    cur  = dict_cursor(conn)
    try:
        cur.execute("SELECT paper_planned FROM jobs WHERE id = %s", [job_id])
        existing = cur.fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Job not found.")
        new_value = not existing["paper_planned"]
        cur.execute(
            "UPDATE jobs SET paper_planned = %s, updated_at = NOW() WHERE id = %s RETURNING *",
            [new_value, job_id],
        )
        row = dict(cur.fetchone())
        conn.commit()
        _enrich_one(cur, row)
        return JobResponse(**row)
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ── Single job GET / PUT / DELETE ─────────────────────────────────────────────

@router.get("/{job_id}", response_model=JobResponse)
def get_job(job_id: int):
    conn = get_connection()
    cur  = dict_cursor(conn)
    cur.execute("SELECT * FROM jobs WHERE id = %s", [job_id])
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found.")
    d = dict(row)
    _enrich_one(cur, d)
    conn.close()
    return JobResponse(**d)


@router.put("/{job_id}", response_model=JobResponse)
def update_job(job_id: int, job: JobUpdate):
    conn = get_connection()
    cur  = dict_cursor(conn)
    try:
        cur.execute("SELECT * FROM jobs WHERE id = %s", [job_id])
        existing = cur.fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Job not found.")

        ex = dict(existing)
        merged = {
            "customer_name":  job.customer_name  if job.customer_name  is not None else ex["customer_name"],
            "job_name":       job.job_name        if job.job_name       is not None else ex["job_name"],
            "artworks":       job.artworks        if job.artworks       is not None else ex.get("artworks", ""),
            "length":         job.length          if job.length         is not None else ex["length"],
            "width":          job.width           if job.width          is not None else ex["width"],
            "height":         job.height          if job.height         is not None else ex["height"],
            "gsm":            job.gsm             if job.gsm            is not None else ex["gsm"],
            "paper_quality":  job.paper_quality   if job.paper_quality  is not None else ex["paper_quality"],
            "order_quantity": job.order_quantity  if job.order_quantity is not None else ex["order_quantity"],
            "sheet_length":   job.sheet_length    if job.sheet_length   is not None else ex["sheet_length"],
            "sheet_width":    job.sheet_width     if job.sheet_width    is not None else ex["sheet_width"],
            "ups":            job.ups             if job.ups            is not None else ex["ups"],
            "printing_type":  job.printing_type   if job.printing_type  is not None else (ex.get("printing_type") or "outer"),
            "remarks":        job.remarks         if job.remarks        is not None else ex.get("remarks", ""),
        }
        calc = calculate_job(
            order_quantity=merged["order_quantity"],
            ups=merged["ups"],
            sheet_length=merged["sheet_length"],
            sheet_width=merged["sheet_width"],
            gsm=merged["gsm"],
        )
        cur.execute("""
            UPDATE jobs SET
                customer_name      = %s,
                job_name           = %s,
                artworks           = %s,
                length             = %s,
                width              = %s,
                height             = %s,
                gsm                = %s,
                paper_quality      = %s,
                order_quantity     = %s,
                sheet_length       = %s,
                sheet_width        = %s,
                ups                = %s,
                printing_type      = %s,
                remarks            = %s,
                base_sheets        = %s,
                wastage_percentage = %s,
                final_sheets       = %s,
                total_kg           = %s,
                updated_at         = NOW()
            WHERE id = %s
            RETURNING *
        """, [
            merged["customer_name"], merged["job_name"], merged["artworks"],
            merged["length"], merged["width"], merged["height"],
            merged["gsm"], merged["paper_quality"],
            merged["order_quantity"],
            merged["sheet_length"], merged["sheet_width"],
            merged["ups"], merged["printing_type"], merged["remarks"],
            calc["base_sheets"], calc["wastage_percentage"],
            calc["final_sheets"], calc["total_kg"],
            job_id,
        ])
        row = dict(cur.fetchone())
        conn.commit()
        _enrich_one(cur, row)
        return JobResponse(**row)
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.delete("/{job_id}", status_code=204)
def delete_job(job_id: int):
    conn = get_connection()
    cur  = conn.cursor()
    try:
        cur.execute("SELECT id FROM jobs WHERE id = %s", [job_id])
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Job not found.")
        cur.execute("DELETE FROM jobs WHERE id = %s", [job_id])
        conn.commit()
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
